from openpyxl import load_workbook
from dataclasses import dataclass

@dataclass
class Student:
    ci: str
    nombre: str
    apellidos: str
    pais: str
    provincia: str
    municipio: str
    situcion_academica: str
    estado: str
    direccion:str
    fecha_de_nacimiento:str
    grupo:str
    carrera:str
    facultad:str
    tipo_de_curso: str
    correo:str
    fuente_de_ingreso:str
    origen_academico: str
    regimen_de_estudio:str
    natural_de:str
    telefono:str
    fecha_de_ingreso_a_la_es:str
    estado_civil:str
    organizacion_politica:str
    fecha_de_ingreso_al_ces:str
    fecha_de_matricula:str
    sexo:str
    color_de_piel:str
    tipo_de_estudiante:str
    anno_de_estudio:str
    centro_de_trabajo:str
    nombre_padre:str
    na_padre: str
    nombre_madre:str
    na_madre:str
    tipo_servicio_militar:str
    edad:str


def get_sheet(path):
    workbook = load_workbook(filename=path)
    # 36 columns in total
    sheet = workbook.active
    return sheet

def get_header(sheet):
    headers_row = sheet.iter_rows(min_row=1, max_row=1,values_only=True)
    for x in headers_row:    
        return x

def get_data(sheet):
    students = []
    for row in sheet.iter_rows(min_row=2,values_only = True):
        student = Student(  ci= row[0],
                            nombre= row[1],
                            apellidos=row[2],
                            pais=row[3],
                            provincia=row[4],
                            municipio=row[5],
                            situcion_academica=row[6],
                            estado=row[7],
                            direccion=row[8],
                            fecha_de_nacimiento=row[9],
                            grupo=row[10],
                            carrera=row[11],
                            facultad=row[12],
                            tipo_de_curso=row[13],
                            correo=row[14],
                            fuente_de_ingreso=row[15],
                            origen_academico=row[16],
                            regimen_de_estudio=row[17],
                            natural_de=row[18],
                            telefono=row[19],
                            fecha_de_ingreso_a_la_es=row[20],
                            estado_civil=row[21],
                            organizacion_politica=row[22],
                            fecha_de_ingreso_al_ces=row[23],
                            fecha_de_matricula=row[24],
                            sexo=row[25],
                            color_de_piel=row[26],
                            tipo_de_estudiante=row[27],
                            anno_de_estudio=row[28],
                            centro_de_trabajo=row[29],
                            nombre_padre=row[30],
                            na_padre=row[31],
                            nombre_madre=row[32],
                            na_madre=row[33],
                            tipo_servicio_militar=row[34],
                            edad=row[35])
        students.append(student)
    return students
