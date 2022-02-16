import enum
import openpyxl 
from dataclasses import dataclass
from read_file import *
from data_selector import *
import os
import errno

def stud_to_tuple(stud):
    a = [
        stud.ci,
        stud.nombre,
        stud.apellidos,
        stud.pais,
        stud.provincia,
        stud.municipio,
        stud.situacion_academica,
        stud.estado,
        stud.direccion,
        stud.fecha_de_nacimiento,
        "", #stud.grupo,
        stud.carrera,
        stud.facultad,
        stud.tipo_de_curso,
        stud.correo,
        stud.fuente_de_ingreso,
        stud.origen_academico,
        stud.regimen_de_estudio,
        stud.natural_de,
        stud.telefono,
        stud.fecha_de_ingreso_a_la_es,
        stud.estado_civil,
        stud.organizacion_politica,
        stud.fecha_de_ingreso_al_ces,
        stud.fecha_de_matricula,
        stud.sexo,
        stud.color_de_piel,
        stud.tipo_de_estudiante,
        stud.anno_de_estudio,
        stud.centro_de_trabajo,
        stud.nombre_padre,
        stud.na_padre,
        stud.nombre_madre,
        stud.na_madre,
        stud.tipo_servicio_militar,
        stud.edad
    ]
    return tuple(a)

def define_data_base(list_):
    database = []
    for item in list_:
        database.append(stud_to_tuple(item))
    return database

def create_sheet(database, headers, id, group):
    wb = openpyxl.Workbook()
    hoja = wb.active
    
    hoja.append(tuple(headers))
    for producto in database:
        hoja.append(producto)

    try:
        os.mkdir('Grupos')
    except OSError as e:
        if e.errno != errno.EEXIST:
            raise
    g = 1 if id < group else 2
    i = id + 1 if id + 1 <= group else id - group + 1
    wb.save(f'Grupos/Grupo-C1{g}{i}.xlsx')



def create_enrollment(database, id, group, curso):
    wb = openpyxl.Workbook()
    hoja = wb.active

    g = 1 if id < group else 2
    i = id + 1 if id + 1 <= group else id - group + 1

    hoja.append(("UNIVERSIDAD DE LA HABANA",))
    hoja.append(("FACULTAD DE MATEMáTICA Y COMPUTACIóN",))
    hoja.append((f"LISTA DE MATRĺCULA DEL GRUPO DOCENTE. CURSO ESCOLAR {curso}-{curso + 1}",))
    hoja.append((f"Carrera: Ciencia de la Computación    Año: 1ero.   Grupo: C1{g}{i}   Tipo de Curso: CD",))
    hoja.append(("",))
    hoja.append(('No.', "Nombre", 'Apellidos', 'Sexo', 'Situación Acad', 'Fuente de Ingreso'))
    for pos, item in enumerate(database):
        hoja.append((f'{pos+1}', item[1], item[2], item[25][0], "", item[15]))

    try:
        os.mkdir('Matrículas')
    except OSError as e:
        if e.errno != errno.EEXIST:
            raise

    wb.save(f'Matrículas/Matrícula-Grupo-C1{g}{i}.xlsx')
